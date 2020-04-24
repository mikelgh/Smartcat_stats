# Version 1.0 / 2020-04-22 Stats proccessing

import os
import tkinter.filedialog as tk

from openpyxl import Workbook
from openpyxl import load_workbook

# default_dir = (r"D:\OneDrive\2.Trans\1.Platts")
default_dir = (r"D:\OneDrive\2.Trans\1.Platts\2020-03 Stats")
os.chdir(default_dir)

# 指定文件，返回文件名：
# filename = tk.askopenfilename(initialdir=default_dir, title="please select a file to process", \
#                               filetypes=[('xlsx', '*.xlsx'), ('All Files', '*')],)
# wb = load_workbook(filename)
# ws = wb['Stats']
# max = ws.max_row
# title = []
# for col in range(1,16):
#     title.append(ws.cell(row=1, column=col).value)
# print(title)


#创建新工作表，名为 Processing
# wsp = wb.create_sheet(title="Processing")
# wb.save(filename)
#新建数据透视表

import pandas as pd
filename = (r"D:\OneDrive\2.Trans\1.Platts\2020-03 Stats\Stats.xlsx")
sExcelFile = filename
df = pd.read_excel(sExcelFile, sheet_name="Stats")
nrows =df.shape[0]
ncols =df.columns.size
# print('Max Rows:'+str(nrows))
# print('Max Columns:'+str(ncols))

# iRow = 0
# for iCol in range(ncols):
#     print(df.iloc[iRow,iCol])
#
# print(df.head)
# print(df.index)
print(df.columns)
# print(df.describe)
df.shape
df = df.sort_values(["Project",'File name'])
# df.to_excel('Stats_pandas.xlsx', sheet_name='Processed')

# gp = df['Total'].astype(float).groupby([df['Project']]).sum()

# 成功生成各项目的字数合计数 2020-04-22
gp = df.groupby(['Project']).sum()
print(gp)
gp.to_excel('Stats_pandas_sum.xlsx', sheet_name='Sum')

# filename = tk.askopenfilename(initialdir=default_dir, title="please select a file to process", \
#                              filetypes=[('xlsx', '*.xlsx'), ('All Files', '*')],)
# 打开结果所在目录：
os.startfile(default_dir)






