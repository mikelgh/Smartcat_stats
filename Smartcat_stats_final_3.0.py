# -*- coding: utf-8 -*-
# @Time    : 2020/04/29
# @Author  : Michael Li
# @Email   : mikelgh@live.com
# @File    : Smartcat_stats_final.py
# @Version : 3.0
# @notes: 解决 适用于任何月份文件夹，自动分出4个regular projects并放在相应工作表+总计行
# @Software: PyCharm
#


import os
import tkinter.filedialog as tk
from typing import Any, Union

from openpyxl import Workbook
from openpyxl import load_workbook
import re
import pandas as pd
from datetime import datetime
from pandas import Series, DataFrame


def Choose_dir():
    try:
        default_dir = r"D:\OneDrive\2.Trans\1.Platts"
        fdir = tk.askdirectory(initialdir=default_dir, title="please select a directory")
        os.chdir(fdir)
        return (fdir)
    except:
        return "error"


def Create_stats_file(dest_filename):
    dest_filename = dest_filename
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stats'
    ws = wb['Stats']

    ws.cell(column=1, row=1).value = "Type"
    ws.cell(column=2, row=1).value = "Project"
    ws.cell(column=3, row=1).value = "Lang"
    ws.cell(column=4, row=1).value = "Analysis file"
    ws.cell(column=5, row=1).value = "File name"
    ws.cell(column=6, row=1).value = "New"
    ws.cell(column=7, row=1).value = "50-74%"
    ws.cell(column=8, row=1).value = "75-84%"
    ws.cell(column=9, row=1).value = "85-94%"
    ws.cell(column=10, row=1).value = "95-99%"
    ws.cell(column=11, row=1).value = "100%"
    ws.cell(column=12, row=1).value = "101%"
    ws.cell(column=13, row=1).value = "102%"
    ws.cell(column=14, row=1).value = "Rep"
    ws.cell(column=15, row=1).value = "CFRep"
    ws.cell(column=16, row=1).value = "Total"
    wb.save(filename=dest_filename)
    return (dest_filename)


def Extract_stats(fdir, dest_filename):
    # os.walk 语法 参考：https://www.runoob.com/python/os-walk.html
    walkTest_tree = os.walk(fdir)
    file_list = []
    for root, dirs, files in os.walk(fdir, topdown=False):
        for name in files:
            file_list.append(os.path.join(root, name))

    result = Filter(file_list)
    # 逐个读取数据：
    count = 0
    for file in result:
        data = []
        wb = load_workbook(file)
        ws = wb['Sheet1']

        max_row = ws.max_row

        for row in range(28, max_row, 16):
            project_name = ws.cell(row=2, column=2).value
            document_name = ws.cell(column=1, row=row).value
            document_name = document_name.split(': ')
            document_name = document_name[1]
            for datarow in range(row + 2, row + 14):
                data.append(ws.cell(column=2, row=datarow).value)
            data.append(document_name)
            data.append(project_name)
        new_data = []
        new_data = split_list(data, n=14)
        # print(new_data)

    for i in range(len(new_data)):
        new_data_out = new_data[i - 1]
        # print(new_data_out)
        count = count + 1
        # print(count)
        # 保存到统计表(下一个空行）
        wb_out = load_workbook(dest_filename)
        ws_out = wb_out['Stats']
        row_number = ws_out.max_row + 1
        for col in range(6, 17):
            ws_out.cell(row=row_number, column=5).value = new_data_out[-2]
            ws_out.cell(row=row_number, column=2).value = new_data_out[-1]
            po = col - 6
            ws_out.cell(row=row_number, column=col).value = new_data_out[po]
        wb_out.save(dest_filename)


def split_list(l, n=64, new=[]):
    if len(l) <= n:
        new.append(l)
        return new
    else:
        new.append(l[:n])
        return split_list(l[n:], n)


def Filter(datalist):
    # 根据列表中的正则表达式搜索数据
    return [val for val in datalist
            if re.findall(r'Statistics for', val)]


def Get_Sum(dest_filename):
    sExcelFile = dest_filename
    df = pd.read_excel(sExcelFile, sheet_name="Stats")
    nrows = df.shape[0]
    ncols = df.columns.size

    df.shape
    df = df.sort_values(["Project", 'File name'])
    gp = df.groupby(['Project']).sum()
    sum_file = 'Stats_sum.xlsx'

    gp.to_excel(sum_file, sheet_name='Sum')
    return (sum_file)


def Result_Reorder(dest_filename):
    sExcelFile = dest_filename
    df = pd.read_excel(sExcelFile, sheet_name="Stats")
    nrows = df.shape[0]
    ncols = df.columns.size
    df.shape
    df = df.sort_values(["Project", 'File name'])
    df.to_excel(dest_filename, sheet_name='Stats')
    return (dest_filename)


def Sort_Sum(dest_filename):
    sExcelFile = dest_filename
    df = pd.read_excel(sExcelFile, sheet_name="Stats")
    nrows = df.shape[0]
    ncols = df.columns.size
    df.shape

    # 按项目名、文件名排序
    df_sort = df.sort_values(["Project", 'File name'])
    # 按项目分类求和
    df_sum = df_sort.groupby(['Project']).sum()
    df_sum_sum = df_sum.groupby(['Type']).sum()
    df_sum = df_sum.append(df_sum_sum)

    df_lithium_sum = df_sort[df_sort['Project'] == "ZH_Lithium Battery"]
    #各列求和并添加到末尾
    df_lithium_sum1 = df_lithium_sum.groupby(['Project']).sum()
    df_lithium_sum = df_lithium_sum.append(df_lithium_sum1)

    df_petchems_sum = df_sort[df_sort['Project'] == "ZH_Petchems"]
    # 各列求和并添加到末尾
    df_petchems_sum1 = df_petchems_sum.groupby(['Project']).sum()
    df_petchems_sum = df_petchems_sum.append(df_petchems_sum1)


    df_smd_sum = df_sort[df_sort['Project'] == "ZH_SteelMarketsDaily"]
    # 各列求和并添加到末尾
    df_smd_sum1 = df_smd_sum.groupby(['Project']).sum()
    df_smd_sum = df_smd_sum.append(df_smd_sum1)

    df_hrcrebar_sum = df_sort[(df_sort['File name'].str.contains("HRC"))|(df_sort['File name'].str.contains("Rebar"))]
    # 各列求和并添加到末尾
    df_hrcrebar_sum1 = df_hrcrebar_sum.groupby(['Project']).sum()
    df_hrcrebar_sum = df_hrcrebar_sum.append(df_hrcrebar_sum1)

    # pandas writer 方法可以同时向一个工作簿写入多个工作表
    writer = pd.ExcelWriter(dest_filename)
    df_sort.to_excel(writer, 'Stats')
    df_sum.to_excel(writer, 'Sum')
    df_lithium_sum.to_excel(writer, 'Lithium')
    df_petchems_sum.to_excel(writer, 'Petchems')
    df_smd_sum.to_excel(writer, 'SMD')
    df_hrcrebar_sum.to_excel(writer, 'HRC & Rebar')
    writer.save()


def main():
    fdir = Choose_dir()

    dest_filename = 'Stats_'+datetime.now().strftime("%Y-%m-%d_%H%M%S")+'.xlsx'
    Create_stats_file(dest_filename)
    Extract_stats(fdir, dest_filename)
    Sort_Sum(dest_filename)
    os.startfile(fdir)
    os.startfile(dest_filename)


if __name__ == '__main__':
    main()
