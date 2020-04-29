# Version 1.0 / 2020-04-22 测试成功  待优化代码

import os
import tkinter.filedialog as tk

from openpyxl import Workbook
from openpyxl import load_workbook

# 初始化 -- 选择统计文件所在文件夹
default_dir = (r"D:\OneDrive\2.Trans\1.Platts\2020-03 Stats")

fdir = tk.askdirectory(initialdir=default_dir, title="please select a directory")
os.chdir(fdir)
# print("Current Working Directory:", end='')
# print(os.getcwd())

# 创建统计文件

dest_filename = 'Stats.xlsx'
wb = Workbook()
ws = wb.active
ws.title = 'Stats'
ws = wb['Stats']

ws.cell(column=1, row=1).value = "Type"
ws.cell(column=2, row=1).value = "Project"
ws.cell(column=3, row=1).value = "Lang"
ws.cell(column=4, row=1).value = "Analysis file"
ws.cell(column=5, row=1).value = "File name"
ws.cell(column=6, row=1).value = "New"
ws.cell(column=7, row=1).value = "50-74%"
ws.cell(column=8, row=1).value = "75-84%"
ws.cell(column=9, row=1).value = "85-94%"
ws.cell(column=10, row=1).value = "95-99%"
ws.cell(column=11, row=1).value = "100%"
ws.cell(column=12, row=1).value = "101%"
ws.cell(column=13, row=1).value = "102%"
ws.cell(column=14, row=1).value = "Rep"
ws.cell(column=15, row=1).value = "CFRep"
ws.cell(column=16, row=1).value = "Total"
wb.save(filename=dest_filename)

# os.walk 语法 参考：https://www.runoob.com/python/os-walk.html
walkTest_tree = os.walk(fdir)
file_list = []
for root, dirs, files in os.walk(fdir, topdown=False):
    for name in files:
        file_list.append(os.path.join(root, name))

# 导入re模块以使用正则表达式
import re


# 声明列表包含科目编号

# 声明过滤功能
def Filter(datalist):
    # 根据列表中的正则表达式搜索数据
    return [val for val in datalist
            if re.findall(r'Statistics for', val)]


# 筛选出符合条件的stats文件：
result = Filter(file_list)


def split_list(l, n=64, new=[]):
    if len(l) <= n:
        new.append(l)
        return new
    else:
        new.append(l[:n])
        return split_list(l[n:], n)


# 逐个读取数据：
count = 0
for file in result:
    data = []
    wb = load_workbook(file)
    ws = wb['Sheet1']

    max_row = ws.max_row


    for row in range(28, max_row, 16):
        project_name = ws.cell(row=2, column=2).value
        document_name = ws.cell(column=1, row=row).value
        document_name = document_name.split(': ')
        document_name = document_name[1]
        for datarow in range(row + 2, row + 14):
            data.append(ws.cell(column=2, row=datarow).value)
        data.append(document_name)
        data.append(project_name)
    new_data = []
    new_data = split_list(data, n=14)
    # print(new_data)

for i in range(len(new_data)):
    new_data_out = new_data[i - 1]
    print(new_data_out)
    count = count + 1
    print(count)
    # 保存到统计表(下一个空行）
    wb_out = load_workbook(dest_filename)
    ws_out = wb_out['Stats']
    row_number = ws_out.max_row + 1
    for col in range(6, 17):
        ws_out.cell(row=row_number, column=5).value = new_data_out[-2]
        ws_out.cell(row=row_number, column=2).value = new_data_out[-1]
        po = col - 6
        ws_out.cell(row=row_number, column=col).value = new_data_out[po]
    wb_out.save(dest_filename)

# 完成后自动打开相关文件 ？ 如何实现？
# print("Finished Successfully")
